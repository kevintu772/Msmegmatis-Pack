import openpyxl

####
## UTR GENERATOR
## Kevin Tu and Kylie Gallagher 7/25/16
## Takes in a spreadsheet with genes' transcription start sites as well as translation start and end sites, generates the UTR's for
## those genes, then adds UTR data to the spreadsheet


#NOTE:wkbk must only have values
wkbkName = "Msmeg_TSSs_for_HS_students_2.xlsx"
sheetName = "TSSs (LC and HC)"
utrFile = "UTRfolds.txt"
wkbk = openpyxl.load_workbook(wkbkName)
sheet = wkbk.get_sheet_by_name(sheetName)

#represents the file holding all info produced by RNAfold
utrFoldingData = ""
utrFolding = open(utrFile, "r")
for line in utrFolding:
    utrFoldingData += line

#represents the msmeg genome
seq= ""
genome = open('msmeg_genome.txt', 'r')
for line in genome:
    seq+=line
seq =seq.replace("\n", "")

#pulls the sequence of a UTR giving its start and stop locations, as well as the strand on which it is located
def pullUTR(start, end, strand):
    #print("start: " + start)
    utrSeq = seq[(start-1):end]
    basecomplement = {'A':'T','C':'G','G':'C','T':'A'}

    #if on minus strand, get the reverse compliment sequence
    if (strand == "-"):
        utrSeq = [basecomplement[base] for base in utrSeq][::-1]
        utrSeq = "".join(utrSeq)
    return utrSeq

output = open("sequence.seq","w")

#using transcription start sites and translation start codons, get the UTR sequence. Note: UTR sequence include 10 coding nt's
def writeSeqFile():

    sheet.cell(row = 1, column = 32).value = "UTR sequence"
    sheet.cell(row = 1, column = 33).value = "UTR end"
    
    stringInput = ""
    for i in range(2, sheet.max_row + 1):
        #print so we know how close to finishing
        print(sheet.max_row - i)

        #only consider UTR's of size greater than 1 but less than 500 nt's, continue looking untill eof
        if ((len(str(sheet.cell(row = i, column = 1).value)) != 0) & (sheet.cell(row = i, column = 24).value <= 500) & (sheet.cell(row = i, column = 24).value > 1)):
            strand = str(sheet.cell(row = i, column = 6).value)
            print(strand)

            #note that minus strand requires "start" of UTR to be the further downstream location (the end of gene rather than the start of gene)
            if (strand == "-"):
                start = sheet.cell(row = i, column = 22).value - 10
                end = sheet.cell(row = i, column = 5).value

                
                utrEnd = start
            else:
                start = sheet.cell(row = i, column = 5).value
                end = sheet.cell(row = i, column = 21).value + 10
                utrEnd = end


            utrSeq = pullUTR(start,end,strand)
            sheet.cell(row = i, column = 32).value = utrSeq
            sheet.cell(row = i, column = 33).value = end

            #format the gene sequences in RNAfold acceptable format 
            stringInput += ">" + str(sheet.cell(row = i, column = 1).value) + "\n" + utrSeq + "\n"

    output.write(stringInput)

    output.close()

#writes the UTR data to a spreadsheet given data from RNAfold
#UTR data includes: structure, free energy, # of nt's bound in the transcription start site, and # of nt's bound in the translation start site
def writeUTRdata():

    #split data by ">", thus splitting it into data per gene
    dataByGene = utrFoldingData.split(">")
    dataByGene = dataByGene[1:]
    workingRow = 2

    #dataColumnStart represents the leftmost column to start outputting data
    dataColumnStart = 34
    headers = ["UTR Sequence", "UTR Fold Data", "UTR Free Energy", "TSS Bound nt's", "Shine Dalgarno Bound nt's"]
    for i in range(5):
        sheet.cell(row = 1, column = dataColumnStart + i).value = headers[i]

    #for each gene in the data, split by line and pull out needed data
    for gene in dataByGene:
    
        splitByLine = gene.split("\n")
        geneName = splitByLine[0]

        geneSequence = splitByLine[1]
        utrFolds = splitByLine[2][:-9]
        utrFreeEnergy = abs(float(splitByLine[2][-7:-1]))*-1
        tssBoundCount = utrFolds[:3].count("(")

        #NOTE: Shine Dalgarno location semi-arbitrarily picked to be between nt's 7 and 12
        sdBoundCount = utrFolds[6:13].count("(")

        data = [geneSequence, utrFolds, utrFreeEnergy, tssBoundCount, sdBoundCount]

        workingRow = getRowByGene(geneName, workingRow)

            
        for i in range(5):
            sheet.cell(row = workingRow, column = dataColumnStart + i).value = data[i]

        workingRow += 1


    wkbk.save(wkbkName)

#finds the first occurance of a given gene in the excel file, starts searching at given start point 
def getRowByGene(gene, start):
    i = start
    printed = False
    while ((i < sheet.max_row) & (str(sheet.cell(row = i, column = 1).value) != gene)):
        i += 1
        print(str(sheet.max_row-i))
        print(gene)
        printed = True
    if (printed == False):
        print(str(sheet.max_row-i))
    return i


#These two functions must be run one at a time (in the order given). Finding a way to automate command line function through python would fix this

#TO USE: UNCOMMENT ONE OF THE BELOW FUNCTIONS
#writeSeqFile()
#writeUTRdata()


genome.close()
            

            
