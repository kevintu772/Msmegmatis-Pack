import openpyxl
import csv
import os

####
## Protein Sequence Generator
## Kevin Tu and Kylie Gallagher 7/25/16
## Takes in a spreadsheet with genes' translation start and end sites, generates aa sequence for those genes using the msmeg genome


#combines genome into an array
st= ""
f = open('msmeg_genome.txt', 'r')
for line in f:
    st+=line
st =st.replace("\n", "")


#turns DNA sequence into RNA   
def complement(DNA, strand):
    basecomplement = {'A':'A','C':'C','G':'G','T':'U'}
    reversebasecomplement = {'A':'U','C':'G','G':'C','T':'A'}  
    if strand == "-":
        letters = [reversebasecomplement[base] for base in DNA] [::-1]
    else:
        letters = [basecomplement[base] for base in DNA]

   # print("DNA: " + str(letters))
        
    return ''.join(letters)

#turns RNA into codons
def codon(com):
    end = len(com) - (len(com) % 3) -1
    codons = [com[i:i+3] for i in range(0, end, 3)]
    return codons

#makes codons into amino acids
def protein(cod):
    codontable = {
    "UUU":"F", "UUC":"F", "UUA":"L", "UUG":"L",
    "UCU":"S", "UCC":"S", "UCA":"S", "UCG":"S",
    "UAU":"Y", "UAC":"Y", "UAA":"", "UAG":"",
    "UGU":"C", "UGC":"C", "UGA":"", "UGG":"W",
    "CUU":"L", "CUC":"L", "CUA":"L", "CUG":"L",
    "CCU":"P", "CCC":"P", "CCA":"P", "CCG":"P",
    "CAU":"H", "CAC":"H", "CAA":"Q", "CAG":"Q",
    "CGU":"R", "CGC":"R", "CGA":"R", "CGG":"R",
    "AUU":"I", "AUC":"I", "AUA":"I", "AUG":"M",
    "ACU":"T", "ACC":"T", "ACA":"T", "ACG":"T",
    "AAU":"N", "AAC":"N", "AAA":"K", "AAG":"K",
    "AGU":"S", "AGC":"S", "AGA":"R", "AGG":"R",
    "GUU":"V", "GUC":"V", "GUA":"V", "GUG":"V",
    "GCU":"A", "GCC":"A", "GCA":"A", "GCG":"A",
    "GAU":"D", "GAC":"D", "GAA":"E", "GAG":"E",
    "GGU":"G", "GGC":"G", "GGA":"G", "GGG":"G",}
    
    sequence = [codontable[codon] for codon in cod]
    
    if ((sequence [0] == "V") | (cod[0] == "UUG")):
        sequence [0] = "M"
    #print("AA: " + str(sequence))
    return ''.join(sequence)

#opens first sheet
workbookName = 'Patric_test.xlsx'

wb = openpyxl.load_workbook(workbookName)
sheet = wb.get_sheet_by_name('Sheet1')



for r in range(2, sheet.max_row + 1):
    if ((sheet.cell(row = r, column = 15).value == "none") & (str(sheet.cell(row = r, column = 2).value) != "tRNA") & (str(sheet.cell(row = r, column = 2).value) != "rRNA") & (str(sheet.cell(row = r, column = 2).value) != "pseudogene")):
        s = sheet.cell(row = r, column = 5).value
        e = sheet.cell(row = r, column = 6).value

    #splits sequence into the correct part
        DNA = st[s-1:e]
        strand = sheet.cell(row = r, column = 7).value
        
        print(sheet.max_row-r)
    #converts the sequences through the functions into amino acids
        sheet.cell(row= r, column= 15).value = protein(codon(complement(DNA, strand)))

wb.save(workbookName)


        
  



               

