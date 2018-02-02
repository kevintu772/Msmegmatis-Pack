import openpyxl


####
## Phyre Parser
## Kevin Tu and Kylie Gallagher 7/25/16
## Takes in a batch data from PHYRE and parses/outputs it to a given spreadsheet
## outputs the top three matches from PHYRE

data = []
workbook = 0
workbookName = ""
sheet = 0
startCol = 0

#loads the excel file and phyre data into memory
def loadData(phyre, excel, sheetName, modCol):

    global data
    global workbook
    global workbookName
    global sheet
    global startCol

    workbookName = excel
    
    f = open(phyre, "r+")

    #load the phyre data, skiping every 21st line (unessecary info)
    s = ""
    linenum = 1
    for line in f:
        if (((linenum-1) % 21 != 0) & (linenum != 1)):
            s+= line
            
        if linenum%21 == 0:
            data.append(s)
            s = ""

        linenum+=1

    #sort the data for efficiency
    data = sorted(data)


    workbook = openpyxl.load_workbook(excel)
    sheet = workbook.get_sheet_by_name(sheetName)

    #convert the given start column into a number
    startCol = ord(modCol)-64


#finds the first occurance of a given gene in the excel file, starts searching at given start point 
def getRowByGene(gene, start):

    global data
    global workbook
    global sheet

    i = start


    while ((i < sheet.max_row) & (str(sheet.cell(row = i, column = 1).value) != gene)):
        i += 1
        print(str(i))
    return i


#parses the phyre data and puts it into the excel sheet
def parseData():

    global data
    global workbookName
    global sheet
    global startCol

    #first write in the headers
    headers = ["PDB Header", "PDB Molecule", "Confidence", "Identicality"]

    #runs 12 times so that the top three phyre matches can be added to excel
    for i in range(12):
        sheet.cell(row = 1, column = startCol + i).value = headers[i%4] + " " + str((i//4)+1)



    
    workingLine = 2

    #split the data into groups by gene
    for group in data:

        tempCol = startCol
        
        print("column: " + str(tempCol))

        #split the gene groups by phyre match
        for line in group.split("\n")[:3]:

            #split the phyre matches by data point
            dataSplit = line.split("|")
            gene = dataSplit[0]

            #get the line of the excel file that will be edited
            workingLine = getRowByGene(gene, workingLine)   

            print("working line: " + str(workingLine))
            
            #get the 5 pieces of info for the phyre match
            rank = dataSplit[3]
            molecule = dataSplit[9][12:]
            moleculeName = dataSplit[10][25:-1]
            confidence = dataSplit[4]
            seqId = dataSplit[5]

            dataInput = [molecule, moleculeName, confidence, seqId]

            #write the info into excel
            for i in range(4):
                sheet.cell(row = workingLine, column = tempCol).value = dataInput[i]

                print("cell: " + str(i))
                tempCol += 1

    workbook.save(workbookName)
            
            
        
            

loadData("phyre_summary.txt", "PATRIC_annotations_msmeg_NC_008596_abbreviated.xlsx", "Sheet1", "P")
parseData()

